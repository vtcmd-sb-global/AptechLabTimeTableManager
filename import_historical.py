#!/usr/bin/env python3
"""Import historical monthly BSR sheets into the database (strict, clean)."""

from __future__ import annotations

import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from database import init_db, db_session

ROOT = Path(__file__).parent
HIST_CANDIDATES = [
    Path("/home/workdir/attachments/21-August-2026 GLS Labstatus.xlsx"),
    ROOT / "data" / "21-August-2026 GLS Labstatus.xlsx",
    Path("/home/workdir/attachments/GLS_Labstatus_template.xlsx"),
]

SLOT_DAY_COLS = {
    ("B", "MWF"): 2, ("B", "TTS"): 8,
    ("C", "MWF"): 14, ("C", "TTS"): 20,
    ("D", "MWF"): 26, ("D", "TTS"): 30,
    ("E", "MWF"): 36, ("E", "TTS"): 42,
    ("F", "MWF"): 48, ("F", "TTS"): 56,
    ("G", "MWF"): 62, ("G", "TTS"): 68,
}

MONTH_NAME_TO_NUM = {
    "january": 1, "february": 2, "feburary": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
}

# Canonical faculty roster (static reference data from BSR)
CANONICAL_FACULTY = [
    ("MS", "Shahbaz", "V"),
    ("AA", "Aman Ali", "V"),
    ("HM", "Hafiz Moiz", "V"),
    ("T", "Talha Qureshi", "P"),
    ("A", "Aousaja", "P"),
    ("BF", "M Bilal Faroqui", "P"),
    ("RA", "Riaz Ahmed", "P"),
    ("MAR", "M Ashar Rehan", "P"),
    ("Z", "Syed Zaid", "V"),
    ("KA", "Kashif Ali", "P"),
    ("AW", "Abdul Wahab", "P"),
    ("H", "Hadiqa", "V"),
    ("AR", "AR", "V"),
    ("SAA", "SAA", "V"),
    ("AM", "AM", "V"),
    ("HA", "HA", "V"),
    ("MA", "MA", "V"),
    ("MAK", "MAK", "V"),
    ("MT", "MT", "V"),
    ("AD", "AD", "V"),
    ("U", "U", "V"),
]

KNOWN_CAREERS = {
    "CPISM", "DISM", "HDSE I", "HDSE II", "ADSE I", "ADSE II",
    "CDMA", "AID", "CS", "STC", "SP", "SEM I", "OST",
    "HDSE1", "HDSE11", "ADSE1", "ADSE11", "HDSE 1", "HDSE 11",
    "ADSE 1", "ADSE 11",
}

CAREER_ALIASES = {
    "HDSE 1": "HDSE I", "HDSE1": "HDSE I", "HDSE 11": "HDSE II", "HDSE11": "HDSE II",
    "ADSE 1": "ADSE I", "ADSE1": "ADSE I", "ADSE 11": "ADSE II", "ADSE11": "ADSE II",
}


def sheet_to_year_month(name: str) -> Optional[str]:
    n = (name or "").strip().lower().replace(" ", "")
    m = re.match(r"([a-z]+)-(\d{2})$", n)
    if not m:
        return None
    mon = MONTH_NAME_TO_NUM.get(m.group(1))
    if not mon:
        return None
    yy = int(m.group(2))
    year = 2000 + yy if yy < 100 else yy
    return f"{year}-{mon:02d}"


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _is_date_like(v) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    s = _cell_str(v)
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return True
    if re.match(r"^\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}$", s):
        return True
    if re.match(r"^\d{1,2}-[A-Z]{3,9}-\d{2}$", s, re.I):
        return True
    return False


def _fmt_date(v) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s.lower().startswith("admission"):
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d-%B-%y"):
        try:
            return datetime.strptime(s[:20], fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _to_int(v) -> int:
    try:
        if v is None or v == "" or v == "-":
            return 0
        return int(float(v))
    except Exception:
        return 0


def _parse_faculty_line(text) -> Tuple[str, str]:
    """Only accept '(COURSE) INIT' patterns e.g. '(AI) MS'."""
    if text is None or _is_date_like(text):
        return "", ""
    t = _cell_str(text)
    if not t or t.lower().startswith("admission"):
        return "", ""
    if re.search(r"\d{4}", t) and "-" in t:
        return "", ""
    # Must be (something) initials
    m = re.match(r"^\(([^)]+)\)\s*([A-Za-z]{1,5})\s*$", t)
    if not m:
        m = re.match(r"^\(([^)]+)\)\s+([A-Za-z]{1,5})\b", t)
    if not m:
        return "", ""
    course, init = m.group(1).strip(), m.group(2).upper()
    # reject if course looks like day group or empty
    if course.upper() in ("MWF", "TTS", "M/W/F", "T/T/S"):
        return "", ""
    return init, course


def _norm_career(text: str) -> str:
    t = _cell_str(text).upper().replace("  ", " ")
    return CAREER_ALIASES.get(t, t)


def _is_career(text) -> bool:
    if text is None or _is_date_like(text):
        return False
    t = _norm_career(_cell_str(text))
    return t in KNOWN_CAREERS or t in CAREER_ALIASES


def _is_batch_code(text) -> bool:
    if text is None or _is_date_like(text):
        return False
    s = _cell_str(text)
    # e.g. AI-202412C2, PMTZ-202405B, DM-202601C
    return bool(re.match(r"^[A-Za-z]{1,6}-\d{4,}", s))


def extract_faculty_from_sheet(ws) -> List[Tuple[str, str, str]]:
    """Read INITIAL / STATUS / Faculty ACCP from bottom faculty table."""
    found = []
    # scan rows 55-90 for INITIAL header then data
    header_row = None
    for r in range(55, 95):
        v = ws.cell(r, 2).value
        if v and str(v).strip().upper() == "INITIAL":
            header_row = r
            break
    if not header_row:
        return found
    for r in range(header_row + 1, header_row + 20):
        init = ws.cell(r, 2).value
        status = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if not init:
            continue
        init_s = str(init).strip().upper()
        if not re.fullmatch(r"[A-Z]{1,5}", init_s):
            continue
        if _is_date_like(init) or _is_date_like(name):
            continue
        name_s = _cell_str(name) or init_s
        st = str(status or "V").strip().upper()[:1]
        if st not in ("P", "V"):
            st = "V"
        found.append((init_s, name_s, st))
    return found


def load_grid(ws, max_r=100, max_c=75) -> Dict[Tuple[int, int], Any]:
    grid = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True), 1):
        for c_idx, v in enumerate(row, 1):
            if v not in (None, ""):
                grid[(r_idx, c_idx)] = v
    return grid



def parse_sheet_allocations(ws, year_month: str, maps: dict, conn) -> List[dict]:
    """Parse by finding faculty header cells, then reading the block below them."""
    grid = load_grid(ws)
    # lab row markers
    lab_at = {}
    for (r, c), v in grid.items():
        if c == 1 and isinstance(v, str) and re.fullmatch(r"L[1-7]", v.strip().upper()):
            lab_at[r] = v.strip().upper()

    def lab_for_row(r: int) -> str:
        best = None
        for lr, code in lab_at.items():
            if lr <= r + 2:  # L marker can be on module row
                if best is None or lr > best[0]:
                    best = (lr, code)
        return best[1] if best else "L1"

    results = []
    known_inits = set(maps["faculty"].keys())

    # Scan all slot columns for faculty pattern cells
    for (slot, day), start_c in SLOT_DAY_COLS.items():
        for r in range(1, 55):
            fac_txt = grid.get((r, start_c))
            initials, _hint = _parse_faculty_line(fac_txt)
            if not initials:
                continue
            if initials not in known_inits and not re.fullmatch(r"[A-Z]{1,5}", initials):
                continue

            # Block rows relative to faculty row
            batch_r, mod_r, date_r = r + 1, r + 2, r + 3
            career_start = r + 4
            batch_raw = grid.get((batch_r, start_c))
            batch = None
            if batch_raw and not _is_date_like(batch_raw):
                bs = _cell_str(batch_raw)
                if re.search(r"[A-Za-z]", bs) and not bs.lower().startswith("admission"):
                    batch = bs

            module = _cell_str(grid.get((mod_r, start_c)))
            if module.lower().startswith("admission") or _is_date_like(grid.get((mod_r, start_c))):
                module = ""
            # If module row is actually an L marker in col A only, module is still this cell
            mod_clean = re.sub(r"\s+\d+$", "", module).strip() if module else ""
            if re.fullmatch(r"L[1-7]", mod_clean.upper()):
                mod_clean = ""

            mod_start = _fmt_date(grid.get((date_r, start_c)))
            admission = False
            if not mod_start and _cell_str(grid.get((date_r, start_c))).lower().startswith("admission"):
                admission = True

            careers_found = []
            for cr in range(career_start, career_start + 4):
                # stop if we hit next faculty header
                if _parse_faculty_line(grid.get((cr, start_c)))[0]:
                    break
                cval = grid.get((cr, start_c))
                if not _is_career(cval):
                    continue
                stud = _to_int(grid.get((cr, start_c + 4)))
                if stud == 0:
                    stud = _to_int(grid.get((cr, start_c + 3)))
                careers_found.append((_norm_career(_cell_str(cval)), stud))

            # batch start: next date after careers or CI5 row nearby
            batch_start = None
            for br in range(career_start, career_start + 6):
                if str(grid.get((br, 1)) or "").strip().upper() in ("CI5", "C15", "CIS"):
                    batch_start = _fmt_date(grid.get((br, start_c)))
                    break
            if not batch_start:
                for br in range(date_r + 1, date_r + 6):
                    if _is_date_like(grid.get((br, start_c))):
                        batch_start = _fmt_date(grid.get((br, start_c)))
                        break

            if not careers_found:
                continue

            lab_code = lab_for_row(r)
            fid = ensure_faculty(conn, initials, maps)
            slot_id = maps["slots"].get(slot)
            day_id = maps["days"].get(day)
            lab_id = maps["labs"].get(lab_code)
            if not all([fid, slot_id, day_id, lab_id]):
                continue

            for career_code, students in careers_found:
                cid = ensure_career(conn, career_code, maps)
                if not cid:
                    continue
                results.append({
                    "faculty_id": fid,
                    "career_id": cid,
                    "lab_id": lab_id,
                    "time_slot_id": slot_id,
                    "day_group_id": day_id,
                    "batch_code": batch,
                    "module_name": mod_clean or None,
                    "students": students,
                    "module_start_date": mod_start,
                    "actual_start_date": batch_start,
                    "notes": None,
                    "is_admission_open": admission,
                    "report_month": year_month,
                    "career_code": career_code,
                    "faculty_initials": initials,
                })
    return results


def parse_sheet_metrics(ws, year_month: str) -> dict:
    metrics = {
        "year_month": year_month,
        "stc_beg": 0, "stc_end": 0, "sp_beg": 0, "sp_end": 0,
        "total_batches": 0, "total_students": 0,
        "mwf_batches": 0, "mwf_students": 0,
        "tts_batches": 0, "tts_students": 0,
        "report_path": None, "notes": "imported from historical Excel",
    }
    rows = list(ws.iter_rows(min_row=55, max_row=100, max_col=75, values_only=True))
    for row in rows:
        for c, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip().lower()

            def right_num():
                for oc in range(c + 1, min(c + 8, len(row))):
                    nv = row[oc]
                    if nv is not None and nv != "":
                        try:
                            return int(float(nv))
                        except Exception:
                            pass
                return 0

            if "stc beg" in s:
                metrics["stc_beg"] = right_num()
            elif "stc end" in s:
                metrics["stc_end"] = right_num()
            elif "sp beg" in s:
                metrics["sp_beg"] = right_num()
            elif "sp end" in s:
                metrics["sp_end"] = right_num()
            elif "total batch mwf" in s:
                metrics["mwf_batches"] = right_num()
            elif "total batch tts" in s:
                metrics["tts_batches"] = right_num()
            elif s in ("total student", "total students"):
                n = right_num()
                if c >= 55 and c <= 62:
                    metrics["mwf_students"] = n
                elif c >= 63:
                    metrics["tts_students"] = n
            elif "stc total" in s:
                metrics["stc_end"] = metrics["stc_end"] or right_num()
    metrics["total_batches"] = metrics["mwf_batches"] + metrics["tts_batches"]
    if not metrics["total_students"]:
        metrics["total_students"] = metrics["mwf_students"] + metrics["tts_students"]
    return metrics


def ensure_faculty(conn, initials: str, maps: dict) -> Optional[int]:
    """Only use faculty already seeded from roster / Excel faculty table."""
    key = (initials or "").upper()
    if not key or not re.fullmatch(r"[A-Z]{1,5}", key):
        return None
    return maps["faculty"].get(key)


def ensure_career(conn, code: str, maps: dict) -> Optional[int]:
    key = _norm_career(code)
    if not key:
        return None
    if key in maps["careers"]:
        return maps["careers"][key]
    is_stc = 1 if key == "STC" else 0
    is_sp = 1 if key == "SP" else 0
    cur = conn.execute(
        "INSERT INTO careers (code, name, is_stc, is_sp, sort_order) VALUES (?, ?, ?, ?, 150)",
        (key, key, is_stc, is_sp),
    )
    maps["careers"][key] = cur.lastrowid
    return cur.lastrowid


def seed_canonical_faculty(conn) -> None:
    existing = {r[0].upper() for r in conn.execute("SELECT initials FROM faculty").fetchall()}
    for init, name, status in CANONICAL_FACULTY:
        if init.upper() in existing:
            # fix name if currently equals initials or looks like a date
            conn.execute(
                """
                UPDATE faculty SET full_name = ?, status = ?, active = 1
                WHERE UPPER(initials) = ?
                  AND (full_name = initials OR full_name GLOB '*[0-9]*' OR full_name = 'XX')
                """,
                (name, status, init.upper()),
            )
            continue
        conn.execute(
            "INSERT INTO faculty (initials, full_name, status, active) VALUES (?, ?, ?, 1)",
            (init.upper(), name, status),
        )


def purge_bad_data(conn) -> None:
    """Remove date-like faculty, XX faculty, and allocations pointing at them."""
    bad_ids = [
        r[0] for r in conn.execute(
            """
            SELECT id FROM faculty WHERE
                initials = 'XX' OR full_name = 'XX'
                OR initials GLOB '*[0-9]*'
                OR full_name GLOB '[0-9]*'
                OR length(initials) > 6
                OR UPPER(initials) IN ('HTML','PHP','JAVA','SQL','SEO','GIT','FSA','MUI','PWD','STC','SP','WORDPRESS')
            """
        ).fetchall()
    ]
    if bad_ids:
        placeholders = ",".join("?" * len(bad_ids))
        conn.execute(f"DELETE FROM allocations WHERE faculty_id IN ({placeholders})", bad_ids)
        conn.execute(f"DELETE FROM faculty WHERE id IN ({placeholders})", bad_ids)

    # also delete batch-code-looking faculty
    bad2 = [
        r[0] for r in conn.execute(
            "SELECT id FROM faculty WHERE initials LIKE '%-%' OR full_name LIKE '%-%'"
        ).fetchall()
    ]
    if bad2:
        ph = ",".join("?" * len(bad2))
        conn.execute(f"DELETE FROM allocations WHERE faculty_id IN ({ph})", bad2)
        conn.execute(f"DELETE FROM faculty WHERE id IN ({ph})", bad2)


def import_historical(path: Optional[Path] = None, replace_all: bool = True) -> dict:
    init_db()
    src = path
    if src is None:
        for c in HIST_CANDIDATES:
            if c.exists():
                src = c
                break
    if src is None or not Path(src).exists():
        raise FileNotFoundError("Historical BSR Excel not found")
    src = Path(src)
    wb = load_workbook(src, data_only=True)

    summary = {"file": str(src), "months": {}, "total_allocations": 0}

    with db_session() as conn:
        if replace_all:
            conn.execute("DELETE FROM allocations")
            purge_bad_data(conn)
        seed_canonical_faculty(conn)

        # merge faculty names from latest sheets
        for sheet_name in reversed(wb.sheetnames):
            if not sheet_to_year_month(sheet_name):
                continue
            for init, name, status in extract_faculty_from_sheet(wb[sheet_name]):
                row = conn.execute(
                    "SELECT id, full_name FROM faculty WHERE UPPER(initials)=?",
                    (init,),
                ).fetchone()
                if row:
                    if row[1] == init or not row[1]:
                        conn.execute(
                            "UPDATE faculty SET full_name=?, status=?, active=1 WHERE id=?",
                            (name, status, row[0]),
                        )
                else:
                    conn.execute(
                        "INSERT INTO faculty (initials, full_name, status, active) VALUES (?,?,?,1)",
                        (init, name, status),
                    )

        maps = {
            "faculty": {r[0].upper(): r[1] for r in conn.execute("SELECT initials, id FROM faculty").fetchall()},
            "labs": {r[0].upper(): r[1] for r in conn.execute("SELECT lab_code, id FROM labs").fetchall()},
            "careers": {r[0].upper(): r[1] for r in conn.execute("SELECT code, id FROM careers").fetchall()},
            "slots": {r[0].upper(): r[1] for r in conn.execute("SELECT code, id FROM time_slots").fetchall()},
            "days": {r[0].upper(): r[1] for r in conn.execute("SELECT code, id FROM day_groups").fetchall()},
        }
        for i in range(1, 8):
            code = f"L{i}"
            if code not in maps["labs"]:
                cur = conn.execute("INSERT INTO labs (lab_code, pcs, active) VALUES (?,20,1)", (code,))
                maps["labs"][code] = cur.lastrowid

        for sheet_name in wb.sheetnames:
            ym = sheet_to_year_month(sheet_name)
            if not ym:
                continue
            ws = wb[sheet_name]
            if replace_all:
                conn.execute("DELETE FROM allocations WHERE report_month=?", (ym,))
            allocs = parse_sheet_allocations(ws, ym, maps, conn)
            for a in allocs:
                conn.execute(
                    """
                    INSERT INTO allocations (
                        faculty_id, career_id, lab_id, time_slot_id, day_group_id,
                        batch_code, course_title, module_name, students,
                        module_start_date, actual_start_date, notes, is_admission_open,
                        report_month, is_active
                    ) VALUES (?,?,?,?,?,?,NULL,?,?,?,?,?,?,?,1)
                    """,
                    (
                        a["faculty_id"], a["career_id"], a["lab_id"],
                        a["time_slot_id"], a["day_group_id"],
                        a.get("batch_code"), a.get("module_name"), a.get("students", 0),
                        a.get("module_start_date"), a.get("actual_start_date"),
                        a.get("notes"), int(a.get("is_admission_open") or 0), ym,
                    ),
                )
            metrics = parse_sheet_metrics(ws, ym)
            if not metrics.get("total_students"):
                metrics["total_students"] = sum(x.get("students", 0) for x in allocs)
            if not metrics.get("total_batches"):
                metrics["total_batches"] = len(allocs)
            stc_stu = sum(x["students"] for x in allocs if x.get("career_code") == "STC")
            sp_stu = sum(x["students"] for x in allocs if x.get("career_code") == "SP")
            metrics["stc_end"] = metrics.get("stc_end") or stc_stu
            metrics["sp_end"] = metrics.get("sp_end") or sp_stu
            metrics["stc_beg"] = metrics.get("stc_beg") or metrics["stc_end"]
            metrics["sp_beg"] = metrics.get("sp_beg") or metrics["sp_end"]

            existing = conn.execute(
                "SELECT id FROM monthly_bsr_metrics WHERE year_month=?", (ym,)
            ).fetchone()
            fields = (
                "stc_beg", "stc_end", "sp_beg", "sp_end",
                "total_batches", "total_students",
                "mwf_batches", "mwf_students", "tts_batches", "tts_students",
                "report_path", "notes",
            )
            vals = [metrics.get(f, 0) for f in fields]
            if existing:
                sets = ", ".join(f"{f}=?" for f in fields)
                conn.execute(
                    f"UPDATE monthly_bsr_metrics SET {sets} WHERE year_month=?",
                    (*vals, ym),
                )
            else:
                cols = ", ".join(["year_month"] + list(fields))
                ph = ", ".join(["?"] * (1 + len(fields)))
                conn.execute(
                    f"INSERT INTO monthly_bsr_metrics ({cols}) VALUES ({ph})",
                    (ym, *vals),
                )

            summary["months"][ym] = {
                "sheet": sheet_name,
                "allocations": len(allocs),
                "students": metrics.get("total_students", 0),
            }
            summary["total_allocations"] += len(allocs)

        # final purge of any residual garbage faculty
        purge_bad_data(conn)
        # remove faculty names that are career codes / modules if no valid name
        conn.execute(
            """
            DELETE FROM faculty WHERE id NOT IN (SELECT DISTINCT faculty_id FROM allocations)
              AND UPPER(initials) IN (
                'HTML','PHP','JAVA','SQL','SEO','GIT','FSA','MUI','PWD','STC','SP',
                'WORDPRESS','CDMA','CPISM','DISM','OST','AI','MWF','TTS','FBDS',
                'MYSQL','DART','VAT','ENT','PBD'
              )
            """
        )

    wb.close()
    return summary


if __name__ == "__main__":
    s = import_historical()
    print("Imported", s["total_allocations"], "allocations")
    for ym, info in sorted(s["months"].items()):
        print(f"  {ym}: {info}")
