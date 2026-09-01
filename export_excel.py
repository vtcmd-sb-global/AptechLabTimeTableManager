#!/usr/bin/env python3
"""
Export Excel reports using the official GLS Labstatus template as the master.

Master template: templates/21-August-2026 GLS Labstatus.xlsx
(contains previous-month BSR sheets that must be preserved).

- The current-month sheet is created or updated from live allocation data.
- All previous-month sheets, formatting, formulas, tables and layouts are kept.
- Sample values on the current-month sheet are cleared; unallocated cells stay blank.
- Allocated blocks receive a unique color per teacher.
- Bottom summary tables are cleared then filled in-place for the current month.
- A Dashboard Summary sheet is (re)created; historical month sheets remain intact.
"""

from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from models import (
    list_allocations, list_labs, list_time_slots, list_day_groups,
    derive_course_from_batch,
)
from calculations import full_dashboard


ROOT = Path(__file__).parent
TEMPLATE_CANDIDATES = [
    ROOT / "templates" / "21-August-2026 GLS Labstatus.xlsx",
    ROOT / "templates" / "GLS_Labstatus_template.xlsx",
    ROOT / "data" / "21-August-2026 GLS Labstatus.xlsx",
]

# Day-group column starts (1-based); each block is 6 columns wide
SLOT_DAY_COLS = {
    ("B", "MWF"): 2,
    ("B", "TTS"): 8,
    ("C", "MWF"): 14,
    ("C", "TTS"): 20,
    ("D", "MWF"): 26,
    ("D", "TTS"): 30,
    ("E", "MWF"): 36,
    ("E", "TTS"): 42,
    ("F", "MWF"): 48,
    ("F", "TTS"): 56,
    ("G", "MWF"): 62,
    ("G", "TTS"): 68,
}

LAB_ROWS = {
    "L1": (4, 10),
    "L2": (11, 17),
    "L3": (18, 25),
    "L4": (26, 32),
    "L5": (33, 39),
    "L6": (40, 46),
    "L7": (47, 54),
}

OFF_FACULTY = 0
OFF_BATCH = 1
OFF_MODULE = 2
OFF_DATE = 3
OFF_CAREER = 4
OFF_EXTRA = 5

FOOTER_SLOT_SUMMARY_ROWS = (55, 56, 58)
FACULTY_HEADER_ROW = 61
FACULTY_FIRST_DATA_ROW = 62
FACULTY_LAST_DATA_ROW = 74
SEMESTER_HEADER_ROW = 86
SEMESTER_FIRST_DATA_ROW = 87
SEMESTER_LAST_DATA_ROW = 97

FAC_COL_INITIAL = 2
FAC_COL_STATUS = 4
FAC_COL_NAME = 5
FAC_COL_LOAD = 17
FAC_COL_CAREER = 20
FAC_COL_STC = 22
FAC_COL_STUDENTS = 23
FAC_COL_TIMINGS = 26

# Faculty × Career matrix on BSR (template cols)
MX_COL_FACULTY = 31
MX_CAREER_COLS = {
    "CPISM": 35, "DISM": 37, "HDSE I": 39, "HDSE1": 39, "HDSE II": 42, "HDSE11": 42,
    "ADSE I": 44, "ADSE1": 44, "ADSE II": 46, "ADSE11": 46,
    "CDMA": 48, "AID": 50, "CS": 51, "STC": 52,
}
MX_COL_TOTAL = 53
MX_TOTAL_ROW = 80

SEM_COL_CODE = 4
SEM_COL_BATCHES = 8
SEM_COL_STUDENTS = 21
SEM_COL_PCT = 26

# Visiting faculty block
VIS_COL_NAME = 31
VIS_COL_NO = 35
VIS_COL_TIME = 39
VIS_FIRST_ROW = 87
VIS_LAST_ROW = 95

# Side totals
SIDE_LABEL_COL = 57
SIDE_VALUE_COL = 67

# Distinct pastel/solid fills – one unique colour per teacher (stable order)
TEACHER_PALETTE = [
    "B2A1C7",  # soft purple
    "9DC3E6",  # soft blue
    "A9D08E",  # soft green
    "F4B183",  # soft orange
    "FFD966",  # soft yellow
    "F4B6C2",  # soft pink
    "8FAADC",  # periwinkle
    "C6E0B4",  # mint
    "D5A6BD",  # mauve
    "BDD7EE",  # light sky
    "FFE699",  # cream yellow
    "C5E0B4",  # light sage
    "F8CBAD",  # peach
    "D0CECE",  # light grey
    "B4C6E7",  # powder blue
    "E2EFDA",  # pale green
    "FFF2CC",  # pale cream
    "DDEBF7",  # ice blue
    "FCE4D6",  # pale peach
    "E2D5F1",  # pale lavender
]

# Header row of each block stays red (template style) when allocated
HEADER_RED = PatternFill(fill_type=FILL_SOLID, fgColor="FF0000")
NO_FILL = PatternFill(fill_type=None)

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2B6CB0")
SECTION_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
SUBHEAD_FILL = PatternFill("solid", fgColor="BEE3F8")
SUBHEAD_FONT = Font(bold=True, name="Calibri", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="C6F6D5")
NORMAL = Font(name="Calibri", size=9)
BOLD = Font(bold=True, name="Calibri", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _faculty_color_map(allocs: list) -> Dict[str, PatternFill]:
    """Stable unique colour per faculty initials (sorted for determinism)."""
    initials = sorted({
        (a.faculty_initials or "").strip().upper()
        for a in allocs
        if (a.faculty_initials or "").strip()
    })
    mapping: Dict[str, PatternFill] = {}
    for i, init in enumerate(initials):
        hex_color = TEACHER_PALETTE[i % len(TEACHER_PALETTE)]
        mapping[init] = PatternFill(fill_type=FILL_SOLID, fgColor=hex_color)
    return mapping


def _safe_set_value(ws, row: int, col: int, value) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def _safe_set_fill(ws, row: int, col: int, fill: PatternFill) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.fill = fill
    return True


def _clear_cell_fill(ws, row: int, col: int) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.fill = NO_FILL
    return True


def _find_template() -> Path:
    for p in TEMPLATE_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("GLS Labstatus template not found under templates/")


def _style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = THIN


def _style_cell(cell, font=NORMAL, align=CENTER, fill=None):
    cell.font = font
    cell.alignment = align
    cell.border = THIN
    if fill:
        cell.fill = fill


def _year_month_to_sheet_name(year_month: str) -> str:
    """Convert 'YYYY-MM' to template-style sheet name e.g. 'August-26'."""
    dt = datetime.strptime(year_month, "%Y-%m")
    return dt.strftime("%B-%y")


def export_workbook(
    output_path: Optional[Path] = None,
    year_month: Optional[str] = None,
    progress_callback=None,
) -> Path:
    """Generate monthly Batch Status Report (BSR).

    year_month: 'YYYY-MM'. Defaults to current month.
    Saves under Output/BSR_YYYY-MM.xlsx (does not overwrite other months).
    progress_callback: optional callable(pct: int, message: str) for UI progress.

    The master template already contains previous-month BSR sheets. Those
    sheets are preserved exactly; only the current-month sheet is created
    or updated with live allocation data. All historical structure,
    formatting, formulas and data remain intact.
    """
    from calculations import build_bsr_month_metrics
    from models import save_monthly_bsr_metrics

    def _prog(pct: int, msg: str = "") -> None:
        if progress_callback:
            try:
                progress_callback(pct, msg)
            except Exception:
                pass

    if year_month is None:
        year_month = datetime.now().strftime("%Y-%m")

    if output_path is None:
        output_path = ROOT / "Output" / f"BSR_{year_month}.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _prog(5, "Loading allocations…")
    allocs = list_allocations(active_only=True, report_month=year_month)
    _prog(12, "Building metrics…")
    metrics = build_bsr_month_metrics(year_month)
    dash = metrics["dashboard"]
    color_map = _faculty_color_map(allocs)

    _prog(20, "Loading master template…")
    template_path = _find_template()
    wb = load_workbook(template_path)

    target_name = _year_month_to_sheet_name(year_month)

    # Locate or create the sheet for the current month
    if target_name in wb.sheetnames:
        lab_ws = wb[target_name]
    else:
        month_sheets = [
            s for s in wb.sheetnames
            if s not in ("Chart1", "Dashboard Summary", "_Data", "BSR")
            and any(ch.isdigit() for ch in s)
        ]
        if month_sheets:
            src = wb[month_sheets[-1]]
            lab_ws = wb.copy_worksheet(src)
            lab_ws.title = target_name
        elif "BSR" in wb.sheetnames:
            lab_ws = wb["BSR"]
            lab_ws.title = target_name
        else:
            lab_ws = wb.create_sheet(target_name)

    _prog(35, "Clearing sample data…")
    _clear_all_sample_data(lab_ws)
    _prog(45, "Writing lab allocations…")
    _fill_lab_data(lab_ws, allocs, color_map)
    _prog(55, "Writing slot summaries…")
    _fill_slot_summaries(lab_ws, allocs)
    _prog(62, "Writing faculty table…")
    _fill_faculty_table(lab_ws, dash)
    _fill_faculty_career_matrix(lab_ws, dash)
    _prog(70, "Writing semester & totals…")
    _fill_visiting_and_day_totals(lab_ws, dash)
    _fill_semester_table(lab_ws, dash)
    _fill_stc_course_table(lab_ws, dash)
    _ensure_summary_formulas(lab_ws)
    _fill_data_sheet(wb, allocs)
    _fill_bsr_month_totals(lab_ws, metrics)

    _prog(80, "Building dashboard…")
    if "Dashboard Summary" in wb.sheetnames:
        del wb["Dashboard Summary"]
    dash_ws = wb.create_sheet("Dashboard Summary", 1)
    _build_dashboard(dash_ws, dash, allocs, color_map)

    for name in list(wb.sheetnames):
        if name.strip().lower() in ("chart", "chart1"):
            del wb[name]

    _prog(90, "Saving workbook…")
    wb.save(output_path)
    _prog(98, "Saving metrics…")

    # Persist monthly snapshot (End-of-Month becomes next month Beginning)
    snap = {k: metrics[k] for k in (
        "stc_beg", "stc_end", "sp_beg", "sp_end",
        "total_batches", "total_students",
        "mwf_batches", "mwf_students", "tts_batches", "tts_students",
    )}
    snap["report_path"] = str(output_path)
    snap["notes"] = f"BSR generated {datetime.now().isoformat(timespec='seconds')}"
    try:
        from models import save_monthly_bsr_metrics
        save_monthly_bsr_metrics(year_month, snap)
    except Exception as e:
        print("Warning: could not save monthly BSR metrics:", e)

    return output_path


def _copy_sheet(src, dst):
    for row in src.iter_rows():
        for cell in row:
            nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = cell.number_format
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)

    for merged in src.merged_cells.ranges:
        try:
            dst.merge_cells(str(merged))
        except Exception:
            pass

    for letter, dim in src.column_dimensions.items():
        dst.column_dimensions[letter].width = dim.width
        if dim.hidden:
            dst.column_dimensions[letter].hidden = True

    for idx, dim in src.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[idx].height = dim.height
        if dim.hidden:
            dst.row_dimensions[idx].hidden = True

    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    if src.freeze_panes:
        dst.freeze_panes = src.freeze_panes
    if src.print_area:
        dst.print_area = src.print_area
    if src.page_setup.orientation:
        dst.page_setup.orientation = src.page_setup.orientation
    if src.page_setup.fitToPage is not None:
        dst.page_setup.fitToPage = src.page_setup.fitToPage
    if src.page_setup.fitToWidth is not None:
        dst.page_setup.fitToWidth = src.page_setup.fitToWidth
    if src.page_setup.fitToHeight is not None:
        dst.page_setup.fitToHeight = src.page_setup.fitToHeight
    if src.sheet_properties.pageSetUpPr is not None:
        dst.sheet_properties.pageSetUpPr.fitToPage = (
            src.sheet_properties.pageSetUpPr.fitToPage
        )


def _clear_all_sample_data(ws):
    """Clear sample values/fills on the lab grid so live data can be written.

    Preserves:
    - All merges, column widths, row heights, fonts from the master template
    - Header rows (time-slot labels on rows 2-3)
    - Row 1 (left blank like historical sheets)
    """
    data_cols = set()
    for start in SLOT_DAY_COLS.values():
        for c in range(start, start + 6):
            data_cols.add(c)

    # 1) Lab grid – clear values + fills (only data columns inside lab blocks)
    for lab_code, (r1, r2) in LAB_ROWS.items():
        for r in range(r1, r2 + 1):
            for c in data_cols:
                _safe_set_value(ws, r, c, None)
                _clear_cell_fill(ws, r, c)

    # 2) Per-slot Career / STC / Total Batch footer
    for r in FOOTER_SLOT_SUMMARY_ROWS:
        for c in data_cols:
            _safe_set_value(ws, r, c, None)
        for start in SLOT_DAY_COLS.values():
            _safe_set_value(ws, r, start, None)

    # 3) Faculty load table + faculty×career matrix data rows
    for r in range(FACULTY_FIRST_DATA_ROW, FACULTY_LAST_DATA_ROW + 1):
        for c in range(1, 75):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue  # keep dynamic formulas
            cell.value = None
    # Matrix total row + side totals residual sample numbers
    for r in range(75, 86):
        for c in range(31, 75):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).startswith("=")):
                cell.value = None
    # Visiting faculty sample names/times + MWF/TTS sample totals
    for r in range(87, 98):
        for c in (31, 35, 39, 45, 49, 56, 60, 64, 67, 68):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, (int, float)) or (isinstance(v, str) and not str(v).strip().endswith(("Month", "MWF", "TTS", "Student", "Students")) and str(v)[:1].isdigit()):
                cell.value = None
            elif isinstance(v, str) and v.strip() and c in (31, 35, 39) and not v.strip().startswith(("Visiting", "Time", "STC", "SP", "Total")):
                # clear sample visiting names/times
                if r >= 87:
                    cell.value = None

    # 4) Semester summary numeric cells
    for r in range(SEMESTER_FIRST_DATA_ROW, SEMESTER_LAST_DATA_ROW + 1):
        for c in (SEM_COL_BATCHES, SEM_COL_STUDENTS, SEM_COL_PCT):
            _safe_set_value(ws, r, c, None)
        for c in range(1, 75):
            if c == SEM_COL_CODE:
                continue
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
                cell.value = None

    # 5) Scattered totals near faculty table
    for r in range(55, 85):
        for c in range(50, 75):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, (int, float)):
                cell.value = None
            elif isinstance(v, str) and v.startswith("="):
                cell.value = None


def _series(a) -> str:
    return (a.course_title or derive_course_from_batch(a.batch_code) or "").strip()


def _slot_code(label_or_code: str) -> str:
    s = (label_or_code or "").strip().upper()
    if not s:
        return ""
    for ch in "BCDEFG":
        if s.startswith(ch):
            return ch
    return s[:1]


def _day_code(code: str) -> str:
    c = (code or "").upper().replace("/", "").replace(" ", "")
    if c in ("MWF", "MW"):
        return "MWF"
    if "T" in c and "S" in c:
        return "TTS"
    if "M" in c and "W" in c:
        return "MWF"
    return code


def _parse_date(val):
    """Return a real datetime (or None) so Excel stores a proper date value
    exactly like the manually-created historical sheets."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    if not s or s.lower() in ("admission open", "n/a", "-"):
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y", "%d-%m-%Y",
                "%Y/%m/%d", "%d %b %Y", "%d %B %Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _fmt_date(val) -> str:
    """String form used only where a text label is required."""
    dt = _parse_date(val)
    if dt is None:
        return ""
    return dt.strftime("%d-%b-%y")


def _set_date_cell(ws, row: int, col: int, val) -> bool:
    """Write a real Excel date (datetime) with the template number format."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    dt = _parse_date(val)
    if dt is None:
        cell.value = None
        return True
    cell.value = dt
    cell.number_format = "d-mmm-yy"
    return True


def _apply_block_fill(ws, r1: int, r2: int, start_c: int, body_fill: PatternFill):
    """Colour the 6-col block: red header row, body_fill on remaining rows."""
    for c in range(start_c, start_c + 6):
        _safe_set_fill(ws, r1 + OFF_FACULTY, c, HEADER_RED)
    for r in range(r1 + OFF_BATCH, r2 + 1):
        for c in range(start_c, start_c + 6):
            _safe_set_fill(ws, r, c, body_fill)



def _ensure_summary_formulas(ws) -> None:
    """Re-apply critical aggregate formulas so the sheet stays dynamic."""
    # Semester %Age
    for r in range(87, 97):
        ws.cell(r, 26).value = f"=IF(U$97=0,0,U{r}/U$97)"
        ws.cell(r, 26).number_format = "0.0%"
    # Semester totals
    ws.cell(97, 8).value = "=SUM(H87:H96)"
    ws.cell(97, 21).value = "=SUM(U87:U96)"
    # Matrix row totals (BA = col 53)
    for r in range(62, 79):
        ws.cell(r, 53).value = f"=SUM(AI{r},AK{r},AM{r},AP{r},AR{r},AT{r},AV{r},AX{r},AY{r},AZ{r})"
    # Matrix column totals (row 80)
    for col, letter in {
        35: "AI", 37: "AK", 39: "AM", 42: "AP", 44: "AR",
        46: "AT", 48: "AV", 50: "AX", 51: "AY", 52: "AZ", 53: "BA",
    }.items():
        ws.cell(80, col).value = f"=SUM({letter}62:{letter}78)"
    # Side linked totals
    ws.cell(83, 67).value = "=BO62"           # CAREER
    ws.cell(86, 67).value = "=BO62+BO84+BO85"  # Total Students



def _fill_bsr_month_totals(ws, metrics: dict) -> None:
    """Write STC/SP beg–end and MWF/TTS monthly totals into template cells."""
    # STC Beg/End of Month (template R87 C45/C49)
    _safe_set_value(ws, 87, 45, metrics.get("stc_beg", 0))
    _safe_set_value(ws, 87, 49, metrics.get("stc_end", 0))
    # SP Beg/End (R90 C45/C49)
    _safe_set_value(ws, 90, 45, metrics.get("sp_beg", 0))
    _safe_set_value(ws, 90, 49, metrics.get("sp_end", 0))
    # MWF / TTS totals (R90)
    _safe_set_value(ws, 90, 56, metrics.get("mwf_batches", 0))
    _safe_set_value(ws, 90, 60, metrics.get("mwf_students", 0))
    _safe_set_value(ws, 90, 64, metrics.get("tts_batches", 0))
    _safe_set_value(ws, 90, 68, metrics.get("tts_students", 0))
    # Side STC TOTAL / SP already handled elsewhere; keep BO85 as stc_end
    _safe_set_value(ws, 85, 67, metrics.get("stc_end", 0))
    _safe_set_value(ws, 84, 67, metrics.get("sp_end", 0))


def _faculty_header_text(series: str, initials: str) -> str:
    """Build faculty header exactly like the manual sheets, e.g. '(AI)   MS'."""
    series = (series or "").strip()
    initials = (initials or "").strip()
    if series and initials:
        # Manual sheets use variable spaces after the series; 3 spaces is the most common look
        return f"({series})   {initials}"
    return initials or series or ""


def _fill_lab_data(ws, allocs: list, color_map: Dict[str, PatternFill]):
    """Write live allocations into the lab grid.

    Admission Open rules
    --------------------
    - Empty slots (no allocation) are left blank – never auto-filled with text.
    - An allocation with is_admission_open=True writes "Admission Open" only in
      the designated faculty-header cell of that slot (template convention).
      Other fields for that allocation may be empty.
    - An allocation with is_admission_open=False writes full batch/module/date
      data and never writes the words "Admission Open".
    """
    grid: Dict[Tuple[str, str, str], list] = defaultdict(list)
    for a in allocs:
        lab = (a.lab_code or "").upper()
        slot = _slot_code(a.time_slot_label or getattr(a, "time_slot_code", "") or "")
        day = _day_code(a.day_group_code or "")
        if lab and slot and day:
            grid[(lab, slot, day)].append(a)

    for lab_code, (r1, r2) in LAB_ROWS.items():
        for (slot, day), start_c in SLOT_DAY_COLS.items():
            items = grid.get((lab_code, slot, day), [])

            if not items:
                # No allocation → leave the block empty (no random text)
                for r in range(r1, r2 + 1):
                    for c in range(start_c, start_c + 6):
                        _safe_set_value(ws, r, c, None)
                        _clear_cell_fill(ws, r, c)
                continue

            # Pure Admission-Open markers vs normal rows
            adm_only = [
                a for a in items
                if getattr(a, "is_admission_open", False)
                and not (a.batch_code or a.module_name or a.module_start_date)
            ]
            normal = [a for a in items if a not in adm_only]

            if adm_only and not normal:
                # Pure Admission Open: only the status text in faculty-header cell
                for r in range(r1, r2 + 1):
                    for c in range(start_c, start_c + 6):
                        _safe_set_value(ws, r, c, None)
                        _clear_cell_fill(ws, r, c)
                _safe_set_value(ws, r1 + OFF_FACULTY, start_c, "Admission Open")
                continue

            # Normal (or mixed) allocation – write full block
            a = normal[0] if normal else items[0]
            series = _series(a)
            fac_line = _faculty_header_text(series, a.faculty_initials)
            init = (a.faculty_initials or "").strip().upper()
            body_fill = color_map.get(init, PatternFill(fill_type=FILL_SOLID, fgColor="D9D9D9"))

            _apply_block_fill(ws, r1, r2, start_c, body_fill)

            _safe_set_value(ws, r1 + OFF_FACULTY, start_c, fac_line or None)
            _safe_set_value(ws, r1 + OFF_BATCH, start_c, a.batch_code or None)
            _safe_set_value(ws, r1 + OFF_MODULE, start_c, a.module_name or None)

            # If marked Admission Open and no module date → status on date row only
            if getattr(a, "is_admission_open", False) and not a.module_start_date:
                _safe_set_value(ws, r1 + OFF_DATE, start_c, "Admission Open")
            else:
                _set_date_cell(ws, r1 + OFF_DATE, start_c, a.module_start_date)

            batch_start_val = None
            for it in items:
                batch_start_val = getattr(it, "actual_start_date", None) or getattr(it, "batch_start_date", None)
                if batch_start_val:
                    break
            if not batch_start_val:
                batch_start_val = getattr(a, "module_start_date", None)

            last_row = r2
            career_last = last_row - 1
            max_career_slots = max(0, career_last - (r1 + OFF_CAREER) + 1)

            write_items = normal if normal else items
            for i, item in enumerate(write_items):
                if i >= max_career_slots:
                    break
                crow = r1 + OFF_CAREER + i
                _safe_set_value(ws, crow, start_c, item.career_code or None)
                stu = item.students if item.students else None
                for offset in (4, 3, 5, 2, 1):
                    if _safe_set_value(ws, crow, start_c + offset, stu):
                        break

            if len(write_items) > max_career_slots and max_career_slots > 0:
                last_crow = r1 + OFF_CAREER + max_career_slots - 1
                extra_note = f"+{len(write_items) - max_career_slots} more"
                for offset in (5, 3, 1, 2):
                    if _safe_set_value(ws, last_crow, start_c + offset, extra_note):
                        break

            if batch_start_val:
                _set_date_cell(ws, last_row, start_c, batch_start_val)
            elif getattr(a, "notes", None) and len(write_items) <= 1:
                _safe_set_value(ws, last_row, start_c, str(a.notes)[:24])




def _fill_stc_course_table(ws, dash: dict):
    """Fill STC short-course rows from dynamic breakdown (career=STC, by module)."""
    breakdown = list(dash.get("stc_module_breakdown") or [])
    total_stc = sum(int(i.get("students") or 0) for i in breakdown)

    for r in range(63, 76):
        for c in (58, 59, 67):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and str(cell.value).startswith("="):
                continue
            cell.value = None
    _safe_set_value(ws, 63, 57, "STC:")

    for i, item in enumerate(breakdown[:13]):
        r = 63 + i
        _safe_set_value(ws, r, 58, f"{i + 1}-")
        _safe_set_value(ws, r, 59, item.get("name") or "")
        _safe_set_value(ws, r, 67, int(item.get("students") or 0))
        ws.cell(r, 70).value = f"=IF(BO$85=0,0,BO{r}/BO$85)"
        ws.cell(r, 70).number_format = "0%"

    _safe_set_value(ws, 85, 67, total_stc)  # STC TOTAL



def _fill_slot_summaries(ws, allocs: list):
    career_stu: Dict[Tuple[str, str], int] = defaultdict(int)
    stc_stu: Dict[Tuple[str, str], int] = defaultdict(int)
    batches: Dict[Tuple[str, str], set] = defaultdict(set)

    for a in allocs:
        slot = _slot_code(a.time_slot_label or getattr(a, "time_slot_code", "") or "")
        day = _day_code(a.day_group_code or "")
        if not slot or not day:
            continue
        key = (slot, day)
        code = (a.career_code or "").upper()
        if code in ("STC", "SP") or getattr(a, "is_stc", False):
            stc_stu[key] += a.students or 0
        else:
            career_stu[key] += a.students or 0
        if a.batch_code:
            batches[key].add(a.batch_code)

    for (slot, day), start_c in SLOT_DAY_COLS.items():
        key = (slot, day)
        cs = career_stu.get(key, 0)
        ss = stc_stu.get(key, 0)
        bc = len(batches.get(key, set()))
        if cs or ss or bc:
            _safe_set_value(ws, 55, start_c, "Career")
            _safe_set_value(ws, 55, start_c + 4, cs if cs else None)
            _safe_set_value(ws, 56, start_c, "STC/SP" if ss else "STC")
            _safe_set_value(ws, 56, start_c + 4, ss if ss else None)
            _safe_set_value(ws, 58, start_c, "Total Batch")
            _safe_set_value(ws, 58, start_c + 4, bc if bc else None)


def _fill_faculty_table(ws, dash: dict):
    """Left faculty-load table + base side totals (BO62/BO84/BO85).
    CAREER (BO83) and Total Students (BO86) are Excel formulas.
    """
    rows = [item for item in dash.get("faculty_workload", []) if item.get("load", 0) > 0]
    for i, item in enumerate(rows):
        r = FACULTY_FIRST_DATA_ROW + i
        if r > FACULTY_LAST_DATA_ROW:
            break
        _safe_set_value(ws, r, FAC_COL_INITIAL, item.get("initials") or "")
        _safe_set_value(ws, r, FAC_COL_STATUS, item.get("status") or "")
        _safe_set_value(ws, r, FAC_COL_NAME, item.get("full_name") or "")
        _safe_set_value(ws, r, FAC_COL_LOAD, item.get("load") or 0)
        _safe_set_value(ws, r, FAC_COL_CAREER, item.get("career_summary") or "")
        stc_val = item.get("stc")
        _safe_set_value(ws, r, FAC_COL_STC, stc_val if stc_val not in (None, 0, "-") else "-")
        _safe_set_value(ws, r, FAC_COL_STUDENTS, item.get("students") or 0)
        _safe_set_value(ws, r, FAC_COL_TIMINGS, item.get("timings") or "")

    career_stu = sum(
        i["students"] for i in dash.get("career_summary", [])
        if i.get("code") not in ("Total", "STC", "SP")
    )
    stc = dash.get("stc_sp", {})
    sp_stu = stc.get("sp_students", 0)
    stc_stu = stc.get("stc_students", 0)

    _safe_set_value(ws, FACULTY_FIRST_DATA_ROW, SIDE_LABEL_COL, "Total  Career Students")
    _safe_set_value(ws, FACULTY_FIRST_DATA_ROW, SIDE_VALUE_COL, career_stu)  # BO62
    _safe_set_value(ws, 84, SIDE_LABEL_COL, "SP")
    _safe_set_value(ws, 84, SIDE_VALUE_COL, sp_stu)   # BO84
    _safe_set_value(ws, 85, SIDE_LABEL_COL, "STC TOTAL")
    _safe_set_value(ws, 85, SIDE_VALUE_COL, stc_stu)  # BO85
    # BO83 (=BO62) and BO86 (=BO62+BO84+BO85) restored by _ensure_summary_formulas


def _fill_faculty_career_matrix(ws, dash: dict):
    """Write faculty names + career counts into the matrix body.
    Row Total (BA) and column totals (row 80) are Excel formulas.
    """
    codes, matrix_rows = dash.get("faculty_career_matrix", ([], []))
    # Clear old matrix body values only (not formulas)
    for r in range(FACULTY_FIRST_DATA_ROW, FACULTY_LAST_DATA_ROW + 1):
        for c in list(MX_CAREER_COLS.values()) + [MX_COL_FACULTY]:
            _safe_set_value(ws, r, c, None)

    for i, mrow in enumerate(matrix_rows):
        r = FACULTY_FIRST_DATA_ROW + i
        if r > FACULTY_LAST_DATA_ROW:
            break
        _safe_set_value(ws, r, MX_COL_FACULTY, mrow.get("faculty") or "")
        for code, col in MX_CAREER_COLS.items():
            n = mrow.get(code, 0) or 0
            if not n:
                alts = {
                    "HDSE1": "HDSE I", "HDSE11": "HDSE II",
                    "ADSE1": "ADSE I", "ADSE11": "ADSE II",
                }
                n = mrow.get(alts.get(code, code), 0) or 0
            if n:
                _safe_set_value(ws, r, col, n)
    _safe_set_value(ws, MX_TOTAL_ROW, MX_COL_FACULTY, "Total")


def _fill_visiting_and_day_totals(ws, dash: dict):
    """Visiting faculty list + MWF/TTS batch/student totals."""
    # Clear visiting rows
    for r in range(VIS_FIRST_ROW, VIS_LAST_ROW + 1):
        for c in (VIS_COL_NAME, VIS_COL_NO, VIS_COL_TIME):
            _safe_set_value(ws, r, c, None)

    visiting = dash.get("visiting_faculty", [])
    for i, v in enumerate(visiting):
        r = VIS_FIRST_ROW + i
        if r > VIS_LAST_ROW:
            break
        _safe_set_value(ws, r, VIS_COL_NAME, v.get("name") or "")
        _safe_set_value(ws, r, VIS_COL_NO, v.get("visiting_no") or 0)
        _safe_set_value(ws, r, VIS_COL_TIME, v.get("time") or "")

    # MWF / TTS totals (template labels at R89, values at R90)
    dgt = dash.get("day_group_totals", {})
    mwf = dgt.get("MWF", {"batches": 0, "students": 0})
    tts = dgt.get("TTS", {"batches": 0, "students": 0})
    _safe_set_value(ws, 90, 56, mwf.get("batches") or 0)
    _safe_set_value(ws, 90, 60, mwf.get("students") or 0)
    _safe_set_value(ws, 90, 64, tts.get("batches") or 0)
    _safe_set_value(ws, 90, 68, tts.get("students") or 0)

    # STC beg/end of month – use current STC student count as both (live snapshot)
    stc_stu = dash.get("stc_sp", {}).get("stc_students", 0)
    _safe_set_value(ws, 87, 45, stc_stu)
    _safe_set_value(ws, 87, 49, stc_stu)
    sp_stu = dash.get("stc_sp", {}).get("sp_students", 0)
    _safe_set_value(ws, 90, 45, sp_stu)
    _safe_set_value(ws, 90, 49, sp_stu)


def _fill_semester_table(ws, dash: dict):
    """Write base batch/student counts only.
    %Age (col Z) and Total row (H97/U97) are Excel formulas in the template.
    """
    label_to_code = {
        "CPISM": "CPISM", "DISM": "DISM",
        "HDSE 1": "HDSE I", "HDSE1": "HDSE I", "HDSE I": "HDSE I",
        "HDSE 11": "HDSE II", "HDSE11": "HDSE II", "HDSE II": "HDSE II",
        "ADSE 1": "ADSE I", "ADSE1": "ADSE I", "ADSE I": "ADSE I",
        "ADSE 11": "ADSE II", "ADSE11": "ADSE II", "ADSE II": "ADSE II",
        "SP": "SP", "CS": "CS", "AID": "AID", "CDMA": "CDMA",
        "STC": "STC", "SEM I": "SEM I", "OST": "OST",
    }
    by_code = {item["code"]: item for item in dash.get("career_summary", [])}

    for r in range(SEMESTER_FIRST_DATA_ROW, SEMESTER_LAST_DATA_ROW + 1):
        label = ws.cell(r, SEM_COL_CODE).value
        if label is None:
            continue
        label_s = str(label).strip()
        if label_s.lower() == "total":
            continue  # H97/U97 are SUM formulas
        code = label_to_code.get(label_s, label_s)
        item = by_code.get(code)
        if item is None:
            for k, v in by_code.items():
                if k.replace(" ", "").upper() == label_s.replace(" ", "").upper():
                    item = v
                    break
        if item is None:
            # still zero out so leftover sample numbers do not remain
            _safe_set_value(ws, r, SEM_COL_BATCHES, 0)
            _safe_set_value(ws, r, SEM_COL_STUDENTS, 0)
            continue
        _safe_set_value(ws, r, SEM_COL_BATCHES, item.get("batches", 0))
        _safe_set_value(ws, r, SEM_COL_STUDENTS, item.get("students", 0))
        # SEM_COL_PCT is formula — never write static percent



def _fill_data_sheet(wb, allocs: list) -> None:
    """Populate hidden _Data sheet (one row per allocation) for transparency / future formulas."""
    if "_Data" not in wb.sheetnames:
        from openpyxl.workbook.workbook import Workbook
        ws = wb.create_sheet("_Data")
        headers = [
            "Lab", "DayGroup", "TimeSlot", "FacultyInitials", "FacultyName", "Status",
            "BatchCode", "Module", "Career", "Students", "ModuleStart", "BatchStart",
            "AdmissionOpen", "Notes",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, value=h)
    else:
        ws = wb["_Data"]
        # clear old rows (keep header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    for i, a in enumerate(allocs, start=2):
        ws.cell(i, 1, value=getattr(a, "lab_code", "") or "")
        ws.cell(i, 2, value=getattr(a, "day_group_code", "") or "")
        ws.cell(i, 3, value=getattr(a, "time_slot_label", "") or getattr(a, "time_slot_code", "") or "")
        ws.cell(i, 4, value=getattr(a, "faculty_initials", "") or "")
        ws.cell(i, 5, value=getattr(a, "faculty_name", "") or "")
        ws.cell(i, 6, value=getattr(a, "faculty_status", "") or "")
        ws.cell(i, 7, value=getattr(a, "batch_code", "") or "")
        ws.cell(i, 8, value=getattr(a, "module_name", "") or "")
        ws.cell(i, 9, value=getattr(a, "career_code", "") or "")
        ws.cell(i, 10, value=getattr(a, "students", 0) or 0)
        ws.cell(i, 11, value=_fmt_date(getattr(a, "module_start_date", None)) or "")
        ws.cell(i, 12, value=_fmt_date(getattr(a, "actual_start_date", None)) or "")
        ao = getattr(a, "is_admission_open", False)
        ws.cell(i, 13, value="Yes" if ao else "No")
        ws.cell(i, 14, value=getattr(a, "notes", "") or "")
    try:
        ws.sheet_state = "hidden"
    except Exception:
        pass


def _build_dashboard(ws, dash: dict, allocs: list, color_map: Dict[str, PatternFill]):
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Aptech Batch Status Report (BSR) – {datetime.now().strftime('%d %B %Y')}"
    ws["A1"].font = Font(bold=True, size=16, color="1A365D", name="Calibri")
    ws.row_dimensions[1].height = 28

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="CAREER / SEMESTER SUMMARY")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)

    row = 4
    for col, h in enumerate(["Semester", "No. of Batches", "No. of Students", "%Age"], 1):
        _style_header(ws.cell(row=row, column=col, value=h), SUBHEAD_FILL, SUBHEAD_FONT)

    for item in dash["career_summary"]:
        row += 1
        vals = [
            item["code"], item["batches"], item["students"],
            f"{item['pct'] * 100:.1f}%" if item["code"] != "Total" else "",
        ]
        for col, v in enumerate(vals, 1):
            fill = TOTAL_FILL if item["code"] == "Total" else None
            _style_cell(ws.cell(row=row, column=col, value=v),
                        BOLD if item["code"] == "Total" else NORMAL, fill=fill)

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="MWF / TTS TOTALS")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)
    row += 1
    for col, h in enumerate(["Day Group", "Batches", "Students"], 1):
        _style_header(ws.cell(row=row, column=col, value=h), SUBHEAD_FILL, SUBHEAD_FONT)
    for dg, totals in dash["day_group_totals"].items():
        row += 1
        for col, v in enumerate([dg, totals["batches"], totals["students"]], 1):
            _style_cell(ws.cell(row=row, column=col, value=v))

    row += 2
    stc = dash["stc_sp"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="STC / SP")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)
    row += 1
    for label, key in [
        ("STC Batches", "stc_batches"), ("STC Students", "stc_students"),
        ("SP Batches", "sp_batches"), ("SP Students", "sp_students"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=stc[key])
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value="FACULTY WORKLOAD")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)
    row += 1
    headers = ["Color", "Initial", "Status", "Faculty ACCP", "Load", "Career", "STC", "No. Students", "TTS/MWF"]
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=row, column=col, value=h), SUBHEAD_FILL, SUBHEAD_FONT)
    for item in dash["faculty_workload"]:
        row += 1
        init = (item.get("initials") or "").strip().upper()
        fac_fill = color_map.get(init)
        # colour swatch cell
        swatch = ws.cell(row=row, column=1, value="")
        swatch.border = THIN
        if fac_fill and item.get("load", 0) > 0:
            swatch.fill = fac_fill
        vals = [
            item.get("initials"), item.get("status"), item.get("full_name"),
            item.get("load"), item.get("career_summary"), item.get("stc_count", item.get("stc", 0)),
            item.get("students"), item.get("timings"),
        ]
        for col, v in enumerate(vals, 2):
            _style_cell(ws.cell(row=row, column=col, value=v),
                        align=LEFT if col in (4, 6, 9) else CENTER)

    row += 2
    codes, matrix_rows = dash["faculty_career_matrix"]
    cols_needed = max(2 + len(codes), 4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_needed)
    ws.cell(row=row, column=1, value="FACULTY x CAREER DISTRIBUTION")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)
    row += 1
    headers = ["Faculty"] + list(codes) + ["Total"]
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=row, column=col, value=h), SUBHEAD_FILL, SUBHEAD_FONT)
    for mrow in matrix_rows:
        row += 1
        label = mrow.get("full_name") or mrow.get("initials") or mrow.get("faculty") or ""
        vals = [label] + [mrow.get(c, 0) for c in codes] + [mrow.get("total", mrow.get("Total", 0))]
        for col, v in enumerate(vals, 1):
            _style_cell(ws.cell(row=row, column=col, value=v),
                        align=LEFT if col == 1 else CENTER)

    # ---- Allocations detail (previously a separate sheet) ----
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    ws.cell(row=row, column=1, value="ALLOCATIONS DETAIL")
    _style_header(ws.cell(row=row, column=1), SECTION_FILL, SECTION_FONT)
    row += 1
    alloc_headers = [
        "Lab", "Day Group", "Time Slot", "Faculty", "Status",
        "Batch Code", "Module", "Career", "Students",
        "Module Start", "Batch Start", "Admission Open", "Notes",
    ]
    for col, h in enumerate(alloc_headers, 1):
        _style_header(ws.cell(row=row, column=col, value=h), SUBHEAD_FILL, SUBHEAD_FONT)
    header_row = row
    for a in allocs:
        row += 1
        init = (a.faculty_initials or "").strip().upper()
        fac_fill = color_map.get(init)
        vals = [
            a.lab_code, a.day_group_code, a.time_slot_label,
            f"{a.faculty_initials} – {a.faculty_name}", a.faculty_status,
            a.batch_code, a.module_name, a.career_code, a.students,
            a.module_start_date, a.actual_start_date,
            "Yes" if a.is_admission_open else "", a.notes or "",
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c_idx, value=v)
            cell.font = NORMAL
            cell.alignment = LEFT if c_idx in (4, 7, 13) else CENTER
            cell.border = THIN
            if fac_fill and c_idx == 4:
                cell.fill = fac_fill

    widths = [8, 10, 14, 22, 8, 14, 14, 10, 10, 12, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["A"].width = 10  # colour swatch / lab


if __name__ == "__main__":
    path = export_workbook()
    print(f"Wrote: {path}")
