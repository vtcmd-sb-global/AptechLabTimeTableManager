#!/usr/bin/env python3
"""Excel / BSR preview panel — grid view of a generated .xlsx inside the app."""

from __future__ import annotations

import tkinter as tk
from tkinter import ttk
from pathlib import Path
from typing import Optional, Any, Dict, Tuple, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX


def _argb_to_hex(color) -> Optional[str]:
    """Convert openpyxl Color to #RRGGBB, or None if theme/auto/empty."""
    if color is None:
        return None
    try:
        rgb = getattr(color, "rgb", None)
        if rgb and isinstance(rgb, str) and len(rgb) >= 6:
            # ARGB or RGB
            if len(rgb) == 8:
                return f"#{rgb[2:8]}"
            if len(rgb) == 6:
                return f"#{rgb}"
        # Indexed colours (limited set)
        indexed = getattr(color, "indexed", None)
        if indexed is not None and indexed in COLOR_INDEX:
            val = COLOR_INDEX[indexed]
            if isinstance(val, str) and len(val) >= 6:
                return f"#{val[-6:]}"
    except Exception:
        pass
    return None


def _cell_fill_hex(cell) -> Optional[str]:
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None or fill.fill_type == "none":
            return None
        # solid / pattern
        fg = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
        return _argb_to_hex(fg)
    except Exception:
        return None


def _cell_font_info(cell) -> Dict[str, Any]:
    info = {"bold": False, "size": 9, "color": "#1e293b", "name": "Segoe UI"}
    try:
        f = cell.font
        if f is None:
            return info
        info["bold"] = bool(f.bold)
        if f.size:
            # Cap for UI readability
            info["size"] = max(7, min(14, int(float(f.size))))
        c = _argb_to_hex(f.color) if f.color else None
        if c:
            info["color"] = c
        if f.name:
            info["name"] = f.name
    except Exception:
        pass
    return info


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%d-%b-%y")
        except Exception:
            return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


class ExcelPreviewPanel(ttk.Frame):
    """Tab content: sheet selector + scrollable Excel-like grid."""

    MAX_ROWS = 120
    MAX_COLS = 30
    MIN_COL_W = 40
    MAX_COL_W = 220
    DEFAULT_ROW_H = 22

    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self._path: Optional[Path] = None
        self._wb = None
        self._ws = None
        self._merge_map: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        self._covered: set = set()
        self._build()

    def _build(self):
        # Toolbar
        bar = ttk.Frame(self, padding=(6, 4))
        bar.pack(fill=tk.X)

        ttk.Label(bar, text="Sheet:").pack(side=tk.LEFT, padx=(0, 4))
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(
            bar, textvariable=self.sheet_var, width=22, state="readonly"
        )
        self.sheet_combo.pack(side=tk.LEFT, padx=(0, 8))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_changed)

        ttk.Button(bar, text="Reload", command=self._reload).pack(side=tk.LEFT, padx=2)
        ttk.Button(bar, text="Open in Excel", command=self._open_external).pack(side=tk.LEFT, padx=2)

        self.path_lbl = ttk.Label(bar, text="No workbook loaded", style="Muted.TLabel")
        self.path_lbl.pack(side=tk.LEFT, padx=(12, 0))

        self.info_lbl = ttk.Label(bar, text="", style="Muted.TLabel")
        self.info_lbl.pack(side=tk.RIGHT, padx=4)

        # Grid host with scrollbars
        host = ttk.Frame(self)
        host.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        self.canvas = tk.Canvas(host, bg="#ffffff", highlightthickness=1, highlightbackground="#d0d5dd")
        self.vbar = ttk.Scrollbar(host, orient=tk.VERTICAL, command=self.canvas.yview)
        self.hbar = ttk.Scrollbar(host, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vbar.set, xscrollcommand=self.hbar.set)

        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.inner = tk.Frame(self.canvas, bg="#ffffff")
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Mouse-wheel scroll
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_wheel)

        # Empty state
        self.empty_lbl = tk.Label(
            self.inner,
            text="Generate a Monthly BSR to preview it here.\n\n"
                 "After generation, this tab will show the workbook\n"
                 "with sheets, colours, and layout.",
            bg="#ffffff", fg="#64748b", font=("Segoe UI", 11),
            justify=tk.CENTER,
        )
        self.empty_lbl.pack(expand=True, pady=80)

    def _on_inner_configure(self, _e=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # Keep inner at least as wide as canvas for empty state
        try:
            self.canvas.itemconfigure(self._win, width=max(event.width, self.inner.winfo_reqwidth()))
        except Exception:
            pass

    def _on_mousewheel(self, event):
        try:
            delta = -1 if event.delta > 0 else 1
            if event.delta == 0 and getattr(event, "num", None) == 5:
                delta = 1
            elif event.delta == 0 and getattr(event, "num", None) == 4:
                delta = -1
            self.canvas.yview_scroll(delta, "units")
        except Exception:
            pass

    def _on_shift_wheel(self, event):
        try:
            delta = -1 if event.delta > 0 else 1
            self.canvas.xview_scroll(delta, "units")
        except Exception:
            pass

    # ── Public API ──────────────────────────────────────────────
    def load_file(self, path: str | Path) -> None:
        path = Path(path)
        if not path.exists():
            self.path_lbl.configure(text=f"File not found: {path.name}")
            return
        try:
            if self._wb is not None:
                try:
                    self._wb.close()
                except Exception:
                    pass
            self._wb = load_workbook(path, data_only=False, read_only=False)
            self._path = path
            names = list(self._wb.sheetnames)
            self.sheet_combo.configure(values=names)
            # Prefer current month-looking sheet, else first non-helper
            prefer = None
            for n in names:
                low = n.lower()
                if low not in ("dashboard summary", "_data", "chart", "chart1"):
                    prefer = n
                    break
            self.sheet_var.set(prefer or (names[0] if names else ""))
            self.path_lbl.configure(text=str(path.name))
            self._render_sheet()
        except Exception as e:
            self.path_lbl.configure(text=f"Preview error: {e}")
            self._clear_grid()
            self.empty_lbl.configure(text=f"Could not load workbook:\n{e}")
            self.empty_lbl.pack(expand=True, pady=80)

    def _reload(self):
        if self._path:
            self.load_file(self._path)

    def _open_external(self):
        if not self._path or not self._path.exists():
            return
        import sys, os, subprocess
        p = str(self._path.resolve())
        try:
            if sys.platform.startswith("win"):
                os.startfile(p)  # type: ignore
            elif sys.platform == "darwin":
                subprocess.Popen(["open", p])
            else:
                subprocess.Popen(["xdg-open", p])
        except Exception:
            pass

    def _on_sheet_changed(self, _e=None):
        self._render_sheet()

    def _clear_grid(self):
        for w in self.inner.winfo_children():
            w.destroy()
        self.empty_lbl = tk.Label(
            self.inner, text="", bg="#ffffff", fg="#64748b",
            font=("Segoe UI", 11), justify=tk.CENTER,
        )

    def _build_merge_maps(self, ws):
        self._merge_map.clear()
        self._covered.clear()
        for mr in ws.merged_cells.ranges:
            min_r, min_c, max_r, max_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            self._merge_map[(min_r, min_c)] = (min_r, min_c, max_r, max_c)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) != (min_r, min_c):
                        self._covered.add((r, c))

    def _col_width_px(self, ws, col_idx: int) -> int:
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        w = None
        if dim and dim.width:
            w = float(dim.width)
        if not w:
            w = 10.0
        # Approximate Excel width → pixels
        px = int(w * 7.5) + 8
        return max(self.MIN_COL_W, min(self.MAX_COL_W, px))

    def _row_height_px(self, ws, row_idx: int) -> int:
        dim = ws.row_dimensions.get(row_idx)
        if dim and dim.height:
            # Excel points ≈ 1.33 px
            return max(16, min(60, int(float(dim.height) * 1.25)))
        return self.DEFAULT_ROW_H

    def _render_sheet(self):
        self._clear_grid()
        if self._wb is None:
            self.empty_lbl.configure(text="No workbook loaded.")
            self.empty_lbl.pack(expand=True, pady=80)
            return

        name = (self.sheet_var.get() or "").strip()
        if not name or name not in self._wb.sheetnames:
            self.empty_lbl.configure(text="Select a sheet.")
            self.empty_lbl.pack(expand=True, pady=80)
            return

        ws = self._wb[name]
        self._ws = ws
        self._build_merge_maps(ws)

        max_r = min(ws.max_row or 1, self.MAX_ROWS)
        max_c = min(ws.max_column or 1, self.MAX_COLS)

        # Shrink to used area with content
        used_r, used_c = 1, 1
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                if cell.value is not None:
                    used_r = max(used_r, r)
                    used_c = max(used_c, c)
        # Include a little padding
        max_r = min(max_r, used_r + 2)
        max_c = min(max_c, used_c + 1)

        truncated = (ws.max_row or 0) > self.MAX_ROWS or (ws.max_column or 0) > self.MAX_COLS
        self.info_lbl.configure(
            text=f"{max_r}×{max_c} cells" + (" (preview limited)" if truncated else "")
        )

        # Header row (column letters) + row numbers — Excel-like chrome
        grid = tk.Frame(self.inner, bg="#c0c0c0")
        grid.pack(anchor="nw")

        col_widths = [self._col_width_px(ws, c) for c in range(1, max_c + 1)]
        row_num_w = 40

        # Corner + column headers
        corner = tk.Label(
            grid, text="", bg="#e8eaed", width=4,
            relief="ridge", bd=1, font=("Segoe UI", 8),
        )
        corner.grid(row=0, column=0, sticky="nsew")
        for c in range(1, max_c + 1):
            hdr = tk.Label(
                grid, text=get_column_letter(c),
                bg="#e8eaed", fg="#333",
                font=("Segoe UI", 8, "bold"),
                relief="ridge", bd=1,
            )
            hdr.grid(row=0, column=c, sticky="nsew")
            grid.columnconfigure(c, minsize=col_widths[c - 1])

        for r in range(1, max_r + 1):
            rh = self._row_height_px(ws, r)
            rn = tk.Label(
                grid, text=str(r), bg="#e8eaed", fg="#333",
                font=("Segoe UI", 8), relief="ridge", bd=1, width=4,
            )
            rn.grid(row=r, column=0, sticky="nsew")
            grid.rowconfigure(r, minsize=rh)

            for c in range(1, max_c + 1):
                if (r, c) in self._covered:
                    continue

                cell = ws.cell(r, c)
                text = _format_value(cell.value)
                fill = _cell_fill_hex(cell) or "#ffffff"
                fi = _cell_font_info(cell)

                # Merge span
                span = self._merge_map.get((r, c))
                rowspan = colspan = 1
                if span:
                    rowspan = span[2] - span[0] + 1
                    colspan = span[3] - span[1] + 1

                # Alignment
                anchor = "w"
                try:
                    ha = (cell.alignment.horizontal or "").lower() if cell.alignment else ""
                    if ha in ("center", "centercontinuous"):
                        anchor = "center"
                    elif ha == "right":
                        anchor = "e"
                except Exception:
                    pass

                font = (fi["name"], fi["size"], "bold" if fi["bold"] else "normal")
                # Fallback fonts that may not exist
                if fi["name"] not in ("Segoe UI", "Arial", "Calibri", "Tahoma", "Helvetica"):
                    font = ("Segoe UI", fi["size"], "bold" if fi["bold"] else "normal")

                lbl = tk.Label(
                    grid,
                    text=text,
                    bg=fill,
                    fg=fi["color"],
                    font=font,
                    anchor=anchor,
                    justify=tk.LEFT,
                    relief="solid",
                    bd=1,
                    padx=3,
                    pady=1,
                )
                lbl.grid(
                    row=r, column=c,
                    rowspan=rowspan, columnspan=colspan,
                    sticky="nsew",
                )

        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.xview_moveto(0)
        self.canvas.yview_moveto(0)
